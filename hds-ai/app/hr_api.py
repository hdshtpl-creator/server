"""Sổ nhân sự có cấu trúc và API nhập liệu an toàn.

``employees``/``employment_contracts`` là nguồn sự thật duy nhất của khu vực này.
Tuyệt đối không đếm hay suy diễn nhân sự từ bảng ``users`` (đó chỉ là tài khoản
đăng nhập). Router được tạo qua ``build_router`` để dùng đúng dependency xác thực
của ``api.py``.
"""
from __future__ import annotations

import csv
import hashlib
import io
import re
import unicodedata
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import PurePosixPath
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app import db


INTERNAL_ROLES = {"admin", "ban_qt", "truong_bph", "chuyen_vien", "tro_ly"}
HR_MANAGER_ROLES = {"admin", "ban_qt"}
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 80 * 1024 * 1024
MAX_XLSX_ENTRIES = 1_000
MAX_IMPORT_ROWS = 5_000
MAX_IMPORT_COLUMNS = 40
MAX_CELL_CHARS = 4_000
MAX_CONTRACTS_PER_EMPLOYEE = 100
MAX_LIST_LIMIT = 200

EMPLOYEE_ACTIVE_STATUSES = {"active"}
CONTRACT_EFFECTIVE_STATUSES = {"", "active", "valid", "effective"}


class HRImportError(ValueError):
    """Lỗi định dạng tổng thể khiến không thể đọc file nhập."""


@dataclass
class ImportIssue:
    severity: str
    message: str
    row: int | None = None
    field: str | None = None

    def to_dict(self) -> dict:
        return {key: value for key, value in asdict(self).items() if value is not None}


@dataclass
class ContractRecord:
    contract_no: str
    start_date: date | None = None
    end_date: date | None = None
    status: str = "active"
    source_document_id: int | None = None
    source_row: int | None = None


@dataclass
class EmployeeRecord:
    employee_code: str
    full_name: str
    title: str | None = None
    department_id: int | None = None
    department_code: str | None = None
    employment_status: str = "active"
    active: bool = True
    source_document_id: int | None = None
    contracts: list[ContractRecord] = field(default_factory=list)
    source_rows: list[int] = field(default_factory=list)


@dataclass
class ImportBatch:
    filename: str
    source_rows: int
    records: list[EmployeeRecord]
    issues: list[ImportIssue]

    @property
    def errors(self) -> list[ImportIssue]:
        return [issue for issue in self.issues if issue.severity == "error"]

    @property
    def warnings(self) -> list[ImportIssue]:
        return [issue for issue in self.issues if issue.severity == "warning"]


class ContractIn(BaseModel):
    contract_no: str
    start_date: date | None = None
    end_date: date | None = None
    status: str = "active"
    source_document_id: int | None = Field(default=None, gt=0)


class EmployeeIn(BaseModel):
    employee_code: str
    full_name: str
    title: str | None = None
    department_id: int | None = Field(default=None, gt=0)
    employment_status: str = "active"
    active: bool = True
    source_document_id: int | None = Field(default=None, gt=0)
    contracts: list[ContractIn] = Field(default_factory=list)


_HEADER_ALIASES = {
    "employee_code": {
        "employee_code", "employee_id", "staff_code", "ma_nhan_vien", "ma_nv",
        "ma_nhan_su", "ma",
    },
    "full_name": {
        "full_name", "employee_name", "staff_name", "ho_ten", "ten_nhan_vien",
        "ho_va_ten", "ten",
    },
    "title": {"title", "job_title", "position", "chuc_danh", "vi_tri"},
    "department_id": {"department_id", "dept_id", "id_phong_ban"},
    "department_code": {
        "department_code", "dept_code", "ma_phong_ban", "ma_phong", "phong_ban", "phong",
    },
    "employment_status": {
        "employment_status", "employee_status", "staff_status", "trang_thai_nhan_su",
        "trang_thai_nhan_vien", "trang_thai",
    },
    "active": {"active", "is_active", "dang_hoat_dong", "con_lam_viec"},
    "source_document_id": {
        "source_document_id", "employee_source_document_id", "tai_lieu_nguon_id",
        "id_tai_lieu_nguon",
    },
    "contract_no": {
        "contract_no", "contract_number", "so_hop_dong", "ma_hop_dong", "so_hd",
    },
    "contract_start_date": {
        "contract_start_date", "start_date", "ngay_bat_dau", "ngay_hieu_luc",
    },
    "contract_end_date": {
        "contract_end_date", "end_date", "ngay_ket_thuc", "ngay_het_han",
    },
    "contract_status": {
        "contract_status", "trang_thai_hop_dong", "tinh_trang_hop_dong",
    },
    "contract_source_document_id": {
        "contract_source_document_id", "contract_document_id", "hop_dong_nguon_id",
        "id_tai_lieu_hop_dong",
    },
}
_ALIAS_TO_HEADER = {
    alias: canonical for canonical, aliases in _HEADER_ALIASES.items() for alias in aliases
}

_EMPLOYEE_STATUS_ALIASES = {
    "active": "active", "working": "active", "employed": "active",
    "dang_lam_viec": "active", "dang_lam": "active",
    "inactive": "inactive", "tam_nghi": "inactive", "khong_hoat_dong": "inactive",
    "on_leave": "on_leave", "nghi_phep": "on_leave", "dang_nghi_phep": "on_leave",
    "terminated": "terminated", "left": "terminated", "nghi_viec": "terminated",
    "da_nghi_viec": "terminated", "cham_dut": "terminated",
}
_CONTRACT_STATUS_ALIASES = {
    "active": "active", "valid": "valid", "effective": "effective",
    "con_hieu_luc": "effective", "dang_hieu_luc": "effective",
    "expired": "expired", "het_han": "expired",
    "terminated": "terminated", "cham_dut": "terminated", "da_cham_dut": "terminated",
    "cancelled": "cancelled", "canceled": "cancelled", "huy": "cancelled", "da_huy": "cancelled",
    "draft": "draft", "nhap": "draft",
}


def _fold(value: Any) -> str:
    text = str(value or "").strip().lower().replace("đ", "d")
    text = "".join(
        char for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _clean_text(value: Any, maximum: int, field_name: str, required: bool = False) -> str | None:
    if value is None:
        text = ""
    elif isinstance(value, (date, datetime)):
        text = value.isoformat()
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = " ".join(text.split()).strip()
    if required and not text:
        raise ValueError(f"Thiếu {field_name}")
    if len(text) > maximum:
        raise ValueError(f"{field_name} vượt giới hạn {maximum} ký tự")
    return text or None


def _required_text(value: Any, maximum: int, field_name: str) -> str:
    text = _clean_text(value, maximum, field_name, required=True)
    # ``required=True`` đã loại chuỗi rỗng; assert giữ hợp đồng kiểu dữ liệu rõ ràng.
    assert text is not None
    return text


def _employee_code(value: Any) -> str:
    code = _required_text(value, 100, "mã nhân viên").upper()
    if not re.fullmatch(r"[A-Z0-9][A-Z0-9._-]*", code):
        raise ValueError(
            "Mã nhân viên chỉ được gồm chữ không dấu, số, dấu chấm, gạch dưới hoặc gạch ngang"
        )
    return code


def _parse_positive_int(value: Any, field_name: str) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        raise ValueError(f"{field_name} phải là số nguyên dương")
    try:
        number = int(value)
    except (TypeError, ValueError):
        try:
            as_float = float(str(value).strip())
            if not as_float.is_integer():
                raise ValueError
            number = int(as_float)
        except (TypeError, ValueError):
            raise ValueError(f"{field_name} phải là số nguyên dương") from None
    if number <= 0:
        raise ValueError(f"{field_name} phải là số nguyên dương")
    return number


def _parse_bool(value: Any, default: bool) -> bool:
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, bool):
        return value
    token = _fold(value)
    if token in {"1", "true", "yes", "y", "co", "active", "dang_hoat_dong"}:
        return True
    if token in {"0", "false", "no", "n", "khong", "inactive", "khong_hoat_dong"}:
        return False
    raise ValueError("active chỉ nhận có/không, true/false hoặc 1/0")


def _parse_date(value: Any, field_name: str) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name} phải có dạng YYYY-MM-DD hoặc DD/MM/YYYY")


def _employee_status(value: Any) -> str:
    token = _fold(value or "active")
    status = _EMPLOYEE_STATUS_ALIASES.get(token)
    if status is None:
        raise ValueError(
            "Trạng thái nhân sự không hợp lệ; dùng active, inactive, on_leave hoặc terminated"
        )
    return status


def _contract_status(value: Any) -> str:
    token = _fold(value or "active")
    status = _CONTRACT_STATUS_ALIASES.get(token)
    if status is None:
        raise ValueError(
            "Trạng thái hợp đồng không hợp lệ; dùng active/effective/valid, expired, "
            "terminated, cancelled hoặc draft"
        )
    return status


def is_effective_contract(
    status: str | None,
    start_date: date | None,
    end_date: date | None,
    as_of: date | None = None,
) -> bool:
    """Một hợp đồng có hiệu lực khi đúng cả trạng thái và khoảng ngày (hai đầu bao gồm)."""
    target = as_of or date.today()
    normalized = _fold(status or "")
    if normalized not in CONTRACT_EFFECTIVE_STATUSES:
        return False
    if start_date is not None and start_date > target:
        return False
    if end_date is not None and end_date < target:
        return False
    return True


def _safe_filename(filename: str | None) -> tuple[str, str]:
    name = (filename or "").strip()
    if not name:
        raise HRImportError("File tải lên không có tên")
    if "\x00" in name or "/" in name or "\\" in name or name in {".", ".."}:
        raise HRImportError("Tên file không an toàn; chỉ gửi tên file, không gửi đường dẫn")
    suffix = PurePosixPath(name).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HRImportError("Chỉ hỗ trợ file .csv hoặc .xlsx (không hỗ trợ .xls)")
    if len(name) > 255:
        raise HRImportError("Tên file vượt giới hạn 255 ký tự")
    return name, suffix


def _check_cell(value: Any, row_number: int, column_number: int) -> None:
    if value is not None and len(str(value)) > MAX_CELL_CHARS:
        raise HRImportError(
            f"Ô dòng {row_number}, cột {column_number} vượt giới hạn {MAX_CELL_CHARS} ký tự"
        )


def _read_csv(data: bytes) -> list[tuple[int, list[Any]]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = data.decode("cp1258")
        except UnicodeDecodeError as exc:
            raise HRImportError("CSV phải dùng UTF-8 hoặc Windows-1258") from exc
    if "\x00" in text:
        raise HRImportError("CSV chứa ký tự NUL không hợp lệ")
    sample = text[:8_192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        dialect = type("HDSDialect", (csv.excel,), {"delimiter": delimiter})
    result: list[tuple[int, list[Any]]] = []
    try:
        for row_number, row in enumerate(csv.reader(io.StringIO(text), dialect), start=1):
            if len(row) > MAX_IMPORT_COLUMNS:
                raise HRImportError(
                    f"Dòng {row_number} có quá {MAX_IMPORT_COLUMNS} cột"
                )
            for column_number, value in enumerate(row, start=1):
                _check_cell(value, row_number, column_number)
            result.append((row_number, row))
    except csv.Error as exc:
        raise HRImportError(f"CSV không hợp lệ: {exc}") from exc
    return result


def _check_xlsx_archive(data: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ENTRIES:
                raise HRImportError("File XLSX có quá nhiều thành phần nén")
            if sum(item.file_size for item in entries) > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise HRImportError("File XLSX nở ra quá lớn; hãy chia nhỏ file")
            for item in entries:
                path = PurePosixPath(item.filename)
                if path.is_absolute() or ".." in path.parts:
                    raise HRImportError("File XLSX chứa đường dẫn nén không an toàn")
    except zipfile.BadZipFile as exc:
        raise HRImportError("File .xlsx bị hỏng hoặc không đúng định dạng Excel") from exc


def _read_xlsx(data: bytes) -> list[tuple[int, list[Any]]]:
    _check_xlsx_archive(data)
    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise HRImportError("Không đọc được file XLSX; file có thể bị hỏng") from exc
    try:
        if not workbook.worksheets:
            raise HRImportError("File XLSX không có trang tính")
        sheet = workbook.worksheets[0]
        # Chặn dimension giả cực lớn trước khi openpyxl tạo tuple ô cho từng dòng.
        if sheet.max_column > MAX_IMPORT_COLUMNS:
            raise HRImportError(f"Trang tính có quá {MAX_IMPORT_COLUMNS} cột")
        if sheet.max_row > MAX_IMPORT_ROWS + 1:
            raise HRImportError(f"Trang tính có quá {MAX_IMPORT_ROWS:,} dòng dữ liệu")
        result: list[tuple[int, list[Any]]] = []
        for row_number, cells in enumerate(sheet.iter_rows(), start=1):
            if len(cells) > MAX_IMPORT_COLUMNS:
                raise HRImportError(
                    f"Dòng {row_number} có quá {MAX_IMPORT_COLUMNS} cột"
                )
            values: list[Any] = []
            for column_number, cell in enumerate(cells, start=1):
                if getattr(cell, "data_type", None) == "f":
                    raise HRImportError(
                        f"Ô dòng {row_number}, cột {column_number} chứa công thức; "
                        "hãy dán thành giá trị trước khi nhập"
                    )
                value = getattr(cell, "value", None)
                _check_cell(value, row_number, column_number)
                values.append(value)
            result.append((row_number, values))
        return result
    finally:
        workbook.close()


def _map_headers(raw_headers: list[Any]) -> tuple[list[str | None], list[ImportIssue]]:
    if not raw_headers:
        raise HRImportError("File không có hàng tiêu đề")
    mapped: list[str | None] = []
    used: dict[str, int] = {}
    issues: list[ImportIssue] = []
    for index, value in enumerate(raw_headers, start=1):
        alias = _fold(value)
        canonical = _ALIAS_TO_HEADER.get(alias)
        if canonical and canonical in used:
            raise HRImportError(
                f"Hai cột {used[canonical]} và {index} cùng ánh xạ vào '{canonical}'"
            )
        if canonical:
            used[canonical] = index
        elif alias:
            issues.append(ImportIssue(
                "warning", f"Bỏ qua cột không nhận diện: {value}", row=1,
            ))
        mapped.append(canonical)
    missing = [name for name in ("employee_code", "full_name") if name not in used]
    if missing:
        raise HRImportError(
            "Thiếu cột bắt buộc: " + ", ".join(missing)
        )
    return mapped, issues


def _contract_from_row(values: dict[str, Any], row_number: int) -> ContractRecord | None:
    contract_fields = (
        "contract_no", "contract_start_date", "contract_end_date",
        "contract_status", "contract_source_document_id",
    )
    if not any(values.get(name) not in (None, "") for name in contract_fields):
        return None
    number = _required_text(values.get("contract_no"), 200, "số hợp đồng")
    start = _parse_date(values.get("contract_start_date"), "ngày bắt đầu hợp đồng")
    end = _parse_date(values.get("contract_end_date"), "ngày kết thúc hợp đồng")
    if start and end and start > end:
        raise ValueError("Ngày bắt đầu hợp đồng không được sau ngày kết thúc")
    return ContractRecord(
        contract_no=number.upper(),
        start_date=start,
        end_date=end,
        status=_contract_status(values.get("contract_status")),
        source_document_id=_parse_positive_int(
            values.get("contract_source_document_id"), "ID tài liệu nguồn của hợp đồng"
        ),
        source_row=row_number,
    )


def _employee_from_row(
    values: dict[str, Any], row_number: int, existing: EmployeeRecord | None = None,
) -> EmployeeRecord:
    code = _employee_code(values.get("employee_code"))
    raw_name = _clean_text(values.get("full_name"), 300, "họ tên")
    if raw_name is None and existing is None:
        raise ValueError("Thiếu họ tên")
    if raw_name is None:
        assert existing is not None
        raw_name = existing.full_name
    status = _employee_status(values.get("employment_status"))
    active = _parse_bool(values.get("active"), default=status == "active")
    title = _clean_text(values.get("title"), 300, "chức danh")
    department_code = _clean_text(values.get("department_code"), 100, "mã phòng ban")
    return EmployeeRecord(
        employee_code=code,
        full_name=raw_name,
        title=title,
        department_id=_parse_positive_int(values.get("department_id"), "ID phòng ban"),
        department_code=department_code.upper() if department_code else None,
        employment_status=status,
        active=active,
        source_document_id=_parse_positive_int(
            values.get("source_document_id"), "ID tài liệu nguồn của nhân viên"
        ),
        source_rows=[row_number],
    )


def _same_contract(left: ContractRecord, right: ContractRecord) -> bool:
    return (
        left.contract_no == right.contract_no
        and left.start_date == right.start_date
        and left.end_date == right.end_date
        and left.status == right.status
        and left.source_document_id == right.source_document_id
    )


def _merge_employee(
    existing: EmployeeRecord,
    incoming: EmployeeRecord,
    contract: ContractRecord | None,
    row_number: int,
    issues: list[ImportIssue],
) -> None:
    conflicts = []
    for attr, label in (
        ("full_name", "họ tên"), ("employment_status", "trạng thái nhân sự"),
        ("active", "active"),
    ):
        if getattr(existing, attr) != getattr(incoming, attr):
            conflicts.append(label)
    for attr, label in (
        ("title", "chức danh"), ("department_id", "ID phòng ban"),
        ("department_code", "mã phòng ban"),
        ("source_document_id", "tài liệu nguồn"),
    ):
        old, new = getattr(existing, attr), getattr(incoming, attr)
        if old is None and new is not None:
            setattr(existing, attr, new)
        elif new is not None and old != new:
            conflicts.append(label)
    if conflicts:
        issues.append(ImportIssue(
            "error",
            f"Mã {existing.employee_code} lặp nhưng khác " + ", ".join(conflicts),
            row=row_number,
            field="employee_code",
        ))
        return
    existing.source_rows.append(row_number)
    issues.append(ImportIssue(
        "warning",
        f"Gộp dòng lặp theo mã nhân viên {existing.employee_code}",
        row=row_number,
        field="employee_code",
    ))
    if contract is None:
        return
    old_contract = next(
        (item for item in existing.contracts if item.contract_no == contract.contract_no), None
    )
    if old_contract is None:
        existing.contracts.append(contract)
    elif _same_contract(old_contract, contract):
        issues.append(ImportIssue(
            "warning",
            f"Bỏ qua hợp đồng trùng hoàn toàn {contract.contract_no}",
            row=row_number,
            field="contract_no",
        ))
    else:
        issues.append(ImportIssue(
            "error",
            f"Hợp đồng {contract.contract_no} bị lặp với dữ liệu khác",
            row=row_number,
            field="contract_no",
        ))


def parse_hr_upload(filename: str, data: bytes) -> ImportBatch:
    """Đọc và kiểm tra CSV/XLSX hoàn toàn trong bộ nhớ, không chạm CSDL."""
    safe_name, extension = _safe_filename(filename)
    if not data:
        raise HRImportError("File tải lên rỗng")
    if len(data) > MAX_UPLOAD_BYTES:
        raise HRImportError("File vượt giới hạn 10 MB")
    raw_rows = _read_csv(data) if extension == ".csv" else _read_xlsx(data)
    while raw_rows and not any(value not in (None, "") for value in raw_rows[0][1]):
        raw_rows.pop(0)
    if not raw_rows:
        raise HRImportError("File không có dữ liệu")
    mapped_headers, issues = _map_headers(raw_rows[0][1])
    records: dict[str, EmployeeRecord] = {}
    non_empty_rows = 0
    for row_number, raw_values in raw_rows[1:]:
        if not any(value not in (None, "") for value in raw_values):
            continue
        non_empty_rows += 1
        if non_empty_rows > MAX_IMPORT_ROWS:
            raise HRImportError(f"File vượt giới hạn {MAX_IMPORT_ROWS:,} dòng dữ liệu")
        values = {
            canonical: raw_values[index] if index < len(raw_values) else None
            for index, canonical in enumerate(mapped_headers) if canonical
        }
        code_value = values.get("employee_code")
        try:
            lookup_code = _employee_code(code_value)
        except ValueError:
            lookup_code = None
        existing = records.get(lookup_code) if lookup_code else None
        try:
            employee = _employee_from_row(values, row_number, existing)
            contract = _contract_from_row(values, row_number)
            if contract and len((existing or employee).contracts) >= MAX_CONTRACTS_PER_EMPLOYEE:
                raise ValueError(
                    f"Một nhân viên không được có quá {MAX_CONTRACTS_PER_EMPLOYEE} hợp đồng trong file"
                )
        except ValueError as exc:
            issues.append(ImportIssue("error", str(exc), row=row_number))
            continue
        if existing:
            _merge_employee(existing, employee, contract, row_number, issues)
        else:
            if contract:
                employee.contracts.append(contract)
            records[employee.employee_code] = employee
    if non_empty_rows == 0:
        raise HRImportError("File chỉ có tiêu đề, chưa có dòng dữ liệu")
    return ImportBatch(safe_name, non_empty_rows, list(records.values()), issues)


def _require_internal(user: dict) -> None:
    if user.get("role") not in INTERNAL_ROLES:
        raise HTTPException(403, "Sổ nhân sự chỉ dành cho nhân viên nội bộ")


def _require_manager(user: dict) -> None:
    if user.get("role") not in HR_MANAGER_ROLES:
        raise HTTPException(
            403,
            "Chỉ Admin hoặc Ban quản trị được xem danh tính, hợp đồng và thay đổi sổ nhân sự",
        )


def _record_from_body(body: EmployeeIn) -> EmployeeRecord:
    if len(body.contracts) > MAX_CONTRACTS_PER_EMPLOYEE:
        raise HTTPException(
            422, f"Một nhân viên không được gửi quá {MAX_CONTRACTS_PER_EMPLOYEE} hợp đồng"
        )
    issues: list[str] = []
    try:
        code = _employee_code(body.employee_code)
        name = _required_text(body.full_name, 300, "họ tên")
        title = _clean_text(body.title, 300, "chức danh")
        status = _employee_status(body.employment_status)
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    contracts: list[ContractRecord] = []
    seen_contracts: set[str] = set()
    for index, item in enumerate(body.contracts, start=1):
        try:
            number = _required_text(item.contract_no, 200, "số hợp đồng").upper()
            status_value = _contract_status(item.status)
            if item.start_date and item.end_date and item.start_date > item.end_date:
                raise ValueError("Ngày bắt đầu hợp đồng không được sau ngày kết thúc")
            if number in seen_contracts:
                raise ValueError(f"Số hợp đồng {number} bị lặp trong yêu cầu")
            seen_contracts.add(number)
            contracts.append(ContractRecord(
                number, item.start_date, item.end_date, status_value,
                item.source_document_id,
            ))
        except ValueError as exc:
            issues.append(f"Hợp đồng #{index}: {exc}")
    if issues:
        raise HTTPException(422, {"message": "Dữ liệu hợp đồng không hợp lệ", "errors": issues})
    return EmployeeRecord(
        employee_code=code, full_name=name, title=title,
        department_id=body.department_id, employment_status=status,
        active=body.active, source_document_id=body.source_document_id,
        contracts=contracts,
    )


def _resolve_records(conn, records: list[EmployeeRecord]) -> tuple[list[dict], list[ImportIssue], dict]:
    """Kiểm tra tham chiếu và lập kế hoạch. Không ghi CSDL."""
    issues: list[ImportIssue] = []
    dept_ids = sorted({r.department_id for r in records if r.department_id is not None})
    dept_codes = sorted({r.department_code for r in records if r.department_code})
    document_ids = sorted({
        value
        for record in records
        for value in (
            [record.source_document_id]
            + [contract.source_document_id for contract in record.contracts]
        )
        if value is not None
    })
    with conn.cursor() as cur:
        if dept_ids or dept_codes:
            cur.execute(
                """SELECT id,upper(code) FROM departments
                    WHERE id=ANY(%s) OR upper(code)=ANY(%s)""",
                (dept_ids or [0], dept_codes or [""]),
            )
            dept_rows = cur.fetchall()
        else:
            dept_rows = []
        if document_ids:
            cur.execute(
                """SELECT id,approved,label_verified,coalesce(active,true),
                          coalesce(extraction_status,'ready')
                     FROM documents WHERE id=ANY(%s)""",
                (document_ids,),
            )
            doc_rows = cur.fetchall()
        else:
            doc_rows = []
        codes = [record.employee_code for record in records]
        if codes:
            cur.execute(
                "SELECT id,employee_code FROM employees WHERE upper(employee_code)=ANY(%s)",
                (codes,),
            )
            employee_rows = cur.fetchall()
        else:
            employee_rows = []

    dept_by_id = {row[0]: row[1] for row in dept_rows}
    dept_by_code = {row[1]: row[0] for row in dept_rows}
    docs = {row[0]: row[1:] for row in doc_rows}
    existing_by_code: dict[str, tuple[int, str]] = {}
    for employee_id, stored_code in employee_rows:
        key = stored_code.upper()
        if key in existing_by_code:
            issues.append(ImportIssue(
                "error", f"CSDL có nhiều mã chỉ khác hoa/thường: {stored_code}",
                field="employee_code",
            ))
        existing_by_code[key] = (employee_id, stored_code)

    for document_id in document_ids:
        state = docs.get(document_id)
        if state is None:
            issues.append(ImportIssue(
                "error", f"Không tìm thấy tài liệu nguồn #{document_id}",
                field="source_document_id",
            ))
        elif not state[0] or not state[1] or not state[2] or state[3] != "ready":
            issues.append(ImportIssue(
                "error", f"Tài liệu nguồn #{document_id} chưa được duyệt/sẵn sàng",
                field="source_document_id",
            ))

    resolved: list[dict] = []
    for record in records:
        resolved_department = record.department_id
        row_hint = record.source_rows[0] if record.source_rows else None
        if record.department_id is not None and record.department_id not in dept_by_id:
            issues.append(ImportIssue(
                "error", f"Không tìm thấy phòng ban ID {record.department_id}",
                row=row_hint, field="department_id",
            ))
        if record.department_code:
            code_department = dept_by_code.get(record.department_code)
            if code_department is None:
                issues.append(ImportIssue(
                    "error", f"Không tìm thấy mã phòng ban {record.department_code}",
                    row=row_hint, field="department_code",
                ))
            elif resolved_department is not None and resolved_department != code_department:
                issues.append(ImportIssue(
                    "error", "ID phòng ban và mã phòng ban không cùng một phòng",
                    row=row_hint, field="department_code",
                ))
            else:
                resolved_department = code_department
        existing = existing_by_code.get(record.employee_code)
        resolved.append({
            "employee_code": record.employee_code,
            "full_name": record.full_name,
            "title": record.title,
            "department_id": resolved_department,
            "employment_status": record.employment_status,
            "active": record.active,
            "source_document_id": record.source_document_id,
            "contracts": [asdict(contract) for contract in record.contracts],
            "existing_id": existing[0] if existing else None,
            "source_rows": list(record.source_rows),
        })

    existing_ids = [item["existing_id"] for item in resolved if item["existing_id"]]
    contract_keys: set[tuple[int, str]] = set()
    if existing_ids:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT employee_id,upper(contract_no)
                     FROM employment_contracts
                    WHERE employee_id=ANY(%s) AND contract_no IS NOT NULL""",
                (existing_ids,),
            )
            contract_keys = {(row[0], row[1]) for row in cur.fetchall()}
    employee_create = sum(item["existing_id"] is None for item in resolved)
    employee_update = len(resolved) - employee_create
    contract_create = 0
    contract_update = 0
    for item in resolved:
        for contract in item["contracts"]:
            key = (item["existing_id"], contract["contract_no"])
            if item["existing_id"] and key in contract_keys:
                contract_update += 1
            else:
                contract_create += 1
    plan = {
        "employees_create": employee_create,
        "employees_update": employee_update,
        "contracts_create": contract_create,
        "contracts_update": contract_update,
    }
    return resolved, issues, plan


def _write_records(conn, resolved: list[dict]) -> tuple[list[int], dict]:
    employee_ids: list[int] = []
    created = updated = contracts_created = contracts_updated = 0
    with conn.cursor() as cur:
        for item in resolved:
            employee_id = item["existing_id"]
            if employee_id is None:
                cur.execute(
                    """INSERT INTO employees
                       (employee_code,full_name,title,department_id,employment_status,active,
                        source_document_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (
                        item["employee_code"], item["full_name"], item["title"],
                        item["department_id"], item["employment_status"], item["active"],
                        item["source_document_id"],
                    ),
                )
                employee_id = cur.fetchone()[0]
                created += 1
            else:
                cur.execute(
                    """UPDATE employees
                          SET employee_code=%s,full_name=%s,title=%s,department_id=%s,
                              employment_status=%s,active=%s,source_document_id=%s,
                              updated_at=now()
                        WHERE id=%s""",
                    (
                        item["employee_code"], item["full_name"], item["title"],
                        item["department_id"], item["employment_status"], item["active"],
                        item["source_document_id"], employee_id,
                    ),
                )
                updated += 1
            employee_ids.append(employee_id)
            for contract in item["contracts"]:
                cur.execute(
                    """SELECT id FROM employment_contracts
                        WHERE employee_id=%s AND upper(contract_no)=%s FOR UPDATE""",
                    (employee_id, contract["contract_no"]),
                )
                contract_row = cur.fetchone()
                if contract_row:
                    cur.execute(
                        """UPDATE employment_contracts
                              SET contract_no=%s,start_date=%s,end_date=%s,status=%s,
                                  source_document_id=%s,updated_at=now()
                            WHERE id=%s""",
                        (
                            contract["contract_no"], contract["start_date"],
                            contract["end_date"], contract["status"],
                            contract["source_document_id"], contract_row[0],
                        ),
                    )
                    contracts_updated += 1
                else:
                    cur.execute(
                        """INSERT INTO employment_contracts
                           (employee_id,contract_no,start_date,end_date,status,source_document_id)
                           VALUES (%s,%s,%s,%s,%s,%s)""",
                        (
                            employee_id, contract["contract_no"], contract["start_date"],
                            contract["end_date"], contract["status"],
                            contract["source_document_id"],
                        ),
                    )
                    contracts_created += 1
    return employee_ids, {
        "employees_created": created,
        "employees_updated": updated,
        "contracts_created": contracts_created,
        "contracts_updated": contracts_updated,
    }


def _report(
    batch: ImportBatch,
    sha256: str,
    database_issues: list[ImportIssue],
    plan: dict,
    committed: bool = False,
    result: dict | None = None,
) -> dict:
    all_issues = batch.issues + database_issues
    errors = [item for item in all_issues if item.severity == "error"]
    warnings = [item for item in all_issues if item.severity == "warning"]
    return {
        "ok": not errors,
        "dry_run": not committed,
        "committed": committed,
        "filename": batch.filename,
        "sha256": sha256,
        "rows_read": batch.source_rows,
        "employees_parsed": len(batch.records),
        "contracts_parsed": sum(len(item.contracts) for item in batch.records),
        "error_count": len(errors),
        "warning_count": len(warnings),
        "issues": [item.to_dict() for item in all_issues],
        "plan": plan,
        "result": result,
        "source_of_truth": "employees/employment_contracts",
    }


async def _read_upload(upload: UploadFile) -> tuple[ImportBatch, str]:
    try:
        data = await upload.read(MAX_UPLOAD_BYTES + 1)
    finally:
        await upload.close()
    try:
        batch = parse_hr_upload(upload.filename or "", data)
    except HRImportError as exc:
        raise HTTPException(422, str(exc)) from exc
    return batch, hashlib.sha256(data).hexdigest()


def _employee_detail(employee_id: int) -> dict:
    with db.session(role="internal", admin=True) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT e.id,e.employee_code,e.full_name,e.title,e.department_id,d.code,d.name,
                          e.employment_status,e.active,e.source_document_id,e.created_at,e.updated_at
                     FROM employees e LEFT JOIN departments d ON d.id=e.department_id
                    WHERE e.id=%s""",
                (employee_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Không tìm thấy nhân viên")
            cur.execute(
                """SELECT id,contract_no,start_date,end_date,status,source_document_id,
                          created_at,updated_at
                     FROM employment_contracts WHERE employee_id=%s
                    ORDER BY start_date DESC NULLS LAST,created_at DESC,id DESC""",
                (employee_id,),
            )
            contract_rows = cur.fetchall()
    contracts = [{
        "id": item[0], "contract_no": item[1], "start_date": item[2],
        "end_date": item[3], "status": item[4], "source_document_id": item[5],
        "effective": is_effective_contract(item[4], item[2], item[3]),
        "created_at": item[6], "updated_at": item[7],
    } for item in contract_rows]
    return {
        "id": row[0], "employee_code": row[1], "full_name": row[2], "title": row[3],
        "department_id": row[4], "department_code": row[5], "department_name": row[6],
        "employment_status": row[7], "active": row[8], "source_document_id": row[9],
        "created_at": row[10], "updated_at": row[11], "contracts": contracts,
        "has_effective_contract": any(item["effective"] for item in contracts),
    }


def build_router(current_user) -> APIRouter:
    router = APIRouter(prefix="/hr", tags=["hr"])

    @router.get("/summary")
    def summary(as_of: date | None = None, user=Depends(current_user)):
        _require_internal(user)
        target = as_of or date.today()
        with db.session(role="internal", admin=True) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """WITH effective_employee AS (
                         SELECT DISTINCT employee_id FROM employment_contracts
                          WHERE lower(trim(coalesce(status,''))) IN ('','active','valid','effective')
                            AND (start_date IS NULL OR start_date <= %s)
                            AND (end_date IS NULL OR end_date >= %s)
                       )
                       SELECT count(*),
                              count(*) FILTER (
                                WHERE e.active
                                  AND lower(trim(coalesce(e.employment_status,'')))='active'
                              ),
                              count(*) FILTER (
                                WHERE e.active
                                  AND lower(trim(coalesce(e.employment_status,'')))='active'
                                  AND ee.employee_id IS NOT NULL
                              ),
                              count(*) FILTER (
                                WHERE e.active
                                  AND lower(trim(coalesce(e.employment_status,'')))='active'
                                  AND ee.employee_id IS NULL
                              )
                         FROM employees e
                         LEFT JOIN effective_employee ee ON ee.employee_id=e.id""",
                    (target, target),
                )
                row = cur.fetchone()
                cur.execute(
                    """SELECT count(*),
                              count(*) FILTER (
                                WHERE lower(trim(coalesce(status,''))) IN ('','active','valid','effective')
                                  AND (start_date IS NULL OR start_date <= %s)
                                  AND (end_date IS NULL OR end_date >= %s)
                              ),
                              count(*) FILTER (
                                WHERE lower(trim(coalesce(status,''))) IN ('','active','valid','effective')
                                  AND end_date BETWEEN %s AND %s
                              )
                         FROM employment_contracts""",
                    (target, target, target, target + timedelta(days=30)),
                )
                contracts = cur.fetchone()
        return {
            "as_of": target,
            "total_employee_records": row[0],
            "active_employees": row[1],
            "active_employees_with_effective_contract": row[2],
            "active_employees_without_effective_contract": row[3],
            "total_contract_records": contracts[0],
            "effective_contracts": contracts[1],
            "contracts_expiring_within_30_days": contracts[2],
            "source_of_truth": "employees/employment_contracts",
            "privacy": "Bản tổng hợp không chứa tên hoặc chi tiết hợp đồng",
        }

    @router.get("/employees")
    def employees_list(
        search: str = "", department_id: int | None = None,
        employment_status: str = "", include_inactive: bool = False,
        limit: int = 100, offset: int = 0, user=Depends(current_user),
    ):
        _require_manager(user)
        limit = max(1, min(limit, MAX_LIST_LIMIT))
        offset = max(0, min(offset, 1_000_000))
        search = " ".join(search.split())[:100]
        clauses: list[str] = []
        params: list[Any] = []
        if not include_inactive:
            clauses.append(
                "e.active AND lower(trim(coalesce(e.employment_status,'')))='active'"
            )
        if department_id is not None:
            clauses.append("e.department_id=%s")
            params.append(department_id)
        if employment_status:
            try:
                normalized_status = _employee_status(employment_status)
            except ValueError as exc:
                raise HTTPException(422, str(exc)) from exc
            clauses.append("e.employment_status=%s")
            params.append(normalized_status)
        if search:
            clauses.append("(e.employee_code ILIKE %s OR e.full_name ILIKE %s)")
            params.extend([f"%{search}%", f"%{search}%"])
        where = " WHERE " + " AND ".join(clauses) if clauses else ""
        with db.session(role="internal", admin=True) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT count(*) FROM employees e" + where, tuple(params))
                total = cur.fetchone()[0]
                cur.execute(
                    """SELECT e.id,e.employee_code,e.full_name,e.title,e.department_id,
                              d.code,d.name,e.employment_status,e.active,e.source_document_id,
                              lc.id,lc.contract_no,lc.start_date,lc.end_date,lc.status,
                              lc.source_document_id,lc.effective
                         FROM employees e
                         LEFT JOIN departments d ON d.id=e.department_id
                         LEFT JOIN LATERAL (
                           SELECT c.id,c.contract_no,c.start_date,c.end_date,c.status,
                                  c.source_document_id,
                                  (lower(trim(coalesce(c.status,''))) IN
                                      ('','active','valid','effective')
                                   AND (c.start_date IS NULL OR c.start_date<=current_date)
                                   AND (c.end_date IS NULL OR c.end_date>=current_date)) AS effective
                             FROM employment_contracts c WHERE c.employee_id=e.id
                            ORDER BY effective DESC,c.start_date DESC NULLS LAST,
                                     c.created_at DESC,c.id DESC LIMIT 1
                         ) lc ON true""" + where +
                    " ORDER BY e.full_name,e.employee_code LIMIT %s OFFSET %s",
                    tuple(params + [limit, offset]),
                )
                rows = cur.fetchall()
        items = []
        for row in rows:
            latest = None if row[10] is None else {
                "id": row[10], "contract_no": row[11], "start_date": row[12],
                "end_date": row[13], "status": row[14], "source_document_id": row[15],
                "effective": row[16],
            }
            items.append({
                "id": row[0], "employee_code": row[1], "full_name": row[2],
                "title": row[3], "department_id": row[4], "department_code": row[5],
                "department_name": row[6], "employment_status": row[7],
                "active": row[8], "source_document_id": row[9],
                "current_or_latest_contract": latest,
            })
        return {"total": total, "limit": limit, "offset": offset, "items": items}

    @router.get("/employees/{employee_id}")
    def employee_get(employee_id: int, user=Depends(current_user)):
        _require_manager(user)
        return _employee_detail(employee_id)

    @router.post("/employees", status_code=201)
    def employee_create(body: EmployeeIn, user=Depends(current_user)):
        _require_manager(user)
        record = _record_from_body(body)
        with db.session(role="internal", admin=True) as conn:
            resolved, issues, _ = _resolve_records(conn, [record])
            errors = [item for item in issues if item.severity == "error"]
            if errors:
                raise HTTPException(422, {
                    "message": "Dữ liệu nhân sự không hợp lệ",
                    "errors": [item.to_dict() for item in errors],
                })
            if resolved[0]["existing_id"] is not None:
                raise HTTPException(409, f"Mã nhân viên {record.employee_code} đã tồn tại")
            employee_ids, result = _write_records(conn, resolved)
            db.audit(
                conn, user["id"], "hr_employee_create", "employees", employee_ids[0],
                {"employee_code": record.employee_code, "contract_count": len(record.contracts)},
            )
        return {"ok": True, "employee": _employee_detail(employee_ids[0]), "result": result}

    @router.put("/employees/{employee_code}")
    def employee_upsert(employee_code: str, body: EmployeeIn, user=Depends(current_user)):
        _require_manager(user)
        record = _record_from_body(body)
        try:
            path_code = _employee_code(employee_code)
        except ValueError as exc:
            raise HTTPException(422, str(exc)) from exc
        if path_code != record.employee_code:
            raise HTTPException(422, "Mã nhân viên trên đường dẫn và trong nội dung không khớp")
        with db.session(role="internal", admin=True) as conn:
            resolved, issues, _ = _resolve_records(conn, [record])
            errors = [item for item in issues if item.severity == "error"]
            if errors:
                raise HTTPException(422, {
                    "message": "Dữ liệu nhân sự không hợp lệ",
                    "errors": [item.to_dict() for item in errors],
                })
            was_create = resolved[0]["existing_id"] is None
            employee_ids, result = _write_records(conn, resolved)
            db.audit(
                conn, user["id"], "hr_employee_upsert", "employees", employee_ids[0],
                {
                    "employee_code": record.employee_code, "created": was_create,
                    "contract_count": len(record.contracts),
                },
            )
        return {"ok": True, "created": was_create,
                "employee": _employee_detail(employee_ids[0]), "result": result}

    @router.post("/import/validate")
    async def import_validate(
        file: UploadFile = File(...), user=Depends(current_user),
    ):
        _require_manager(user)
        batch, sha256 = await _read_upload(file)
        with db.session(role="internal", admin=True) as conn:
            _, database_issues, plan = _resolve_records(conn, batch.records)
        return _report(batch, sha256, database_issues, plan)

    @router.post("/import")
    async def import_file(
        file: UploadFile = File(...), dry_run: bool = True, commit: bool = False,
        user=Depends(current_user),
    ):
        _require_manager(user)
        if dry_run and commit:
            raise HTTPException(422, "Không thể vừa dry_run=true vừa commit=true")
        if not dry_run and not commit:
            raise HTTPException(
                422,
                "Để ghi thật, phải xác nhận rõ dry_run=false và commit=true; mặc định chỉ kiểm tra",
            )
        batch, sha256 = await _read_upload(file)
        if dry_run:
            with db.session(role="internal", admin=True) as conn:
                _, database_issues, plan = _resolve_records(conn, batch.records)
            return _report(batch, sha256, database_issues, plan)

        # Một transaction duy nhất: kiểm tra lại tham chiếu, ghi toàn bộ và audit. Bất kỳ
        # lỗi nào (kể cả audit) đều làm db.session rollback toàn bộ lần nhập.
        with db.session(role="internal", admin=True) as conn:
            resolved, database_issues, plan = _resolve_records(conn, batch.records)
            preview = _report(batch, sha256, database_issues, plan)
            if not preview["ok"]:
                raise HTTPException(422, {
                    "message": "File còn lỗi nên chưa ghi bất kỳ dòng nào",
                    "report": preview,
                })
            employee_ids, result = _write_records(conn, resolved)
            db.audit(
                conn, user["id"], "hr_import_commit", "employees", None,
                {
                    "filename": batch.filename, "sha256": sha256,
                    "rows": batch.source_rows, "employees": len(employee_ids), **result,
                },
            )
        return _report(batch, sha256, database_issues, plan, committed=True, result=result)

    return router
